"""
Lead CRM Export — exact DB field names, no renaming.
Three sheets:
  1. Leads       — all scalar/object fields fully flattened (dot-notation)
  2. Obligations — one row per obligation, linked to lead by custom_lead_id
  3. Stats       — counts by status / loan_type / obligation action
"""

import pymongo
from pymongo import MongoClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, sys

# ─── CONFIG ────────────────────────────────────────────────────────────────────
MONGO_URI   = "mongodb://raunak:Raunak%40123@156.67.111.95:27017/admin?authSource=admin"
DB_NAME     = "crm_database"
COLLECTION  = "leads"
OUTPUT_FILE = f"/www/wwwroot/RupiyaMe/leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# ─── HELPERS ───────────────────────────────────────────────────────────────────

def to_str(val):
    """Convert any value to a plain string Excel can handle."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, list):
        # If list of simple scalars → join
        if all(not isinstance(v, (dict, list)) for v in val):
            return ", ".join(to_str(v) for v in val)
        # Otherwise keep as string representation
        return str(val)
    if isinstance(val, dict):
        return str(val)
    return str(val)


def flatten(doc: dict, prefix: str = "", skip_keys: set = None) -> dict:
    """
    Recursively flatten a dict.
    Lists of dicts (e.g. obligations, references) are NOT flattened here —
    they are handled separately.
    Simple lists (strings/numbers) are joined with comma.
    """
    if skip_keys is None:
        skip_keys = set()
    result = {}
    for k, v in doc.items():
        if k in skip_keys:
            continue
        full_key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict):
            result.update(flatten(v, prefix=full_key))
        elif isinstance(v, list):
            if len(v) == 0:
                result[full_key] = ""
            elif all(not isinstance(item, (dict, list)) for item in v):
                # Simple list → join
                result[full_key] = ", ".join(to_str(i) for i in v)
            else:
                # List of objects — skip (handled in separate sheet)
                pass
        else:
            result[full_key] = to_str(v)
    return result


# Columns to SKIP from the main Leads sheet (too large / separate sheet)
SKIP_TOP = {"dynamic_fields", "activity", "field_history", "assignment_history",
             "process_data", "question_responses", "importantquestion"}
SKIP_DYN = {"obligations", "references", "co_applicant_form", "applicant_form"}


def fmt_obligations(obligations: list) -> str:
    """
    Format obligations list as clean readable lines:
      #1 | PL (PERSONAL LOAN) | HDFC BANK | 5,00,000 | 3,20,000 | 8,500 | BT
      #2 | Home Loan         | SBI       | -         | 12,000  | 500   | Obligate
    """
    if not obligations:
        return ""
    lines = []
    for i, ob in enumerate(obligations, start=1):
        if not isinstance(ob, dict):
            continue
        product    = ob.get("product") or "-"
        bank       = ob.get("bank_name") or ob.get("bankName") or "-"
        total_loan = ob.get("total_loan") or ob.get("totalLoan")
        outstanding= ob.get("outstanding")
        emi        = ob.get("emi") or ob.get("foirEmi")
        action     = ob.get("action") or "-"

        def fmt_num(v):
            try:
                n = float(v)
                return f"{n:,.0f}"
            except (TypeError, ValueError):
                return "-"

        line = (f"#{i} | {product} | {bank} | "
                f"{fmt_num(total_loan)} | {fmt_num(outstanding)} | "
                f"{fmt_num(emi)} | {action}")
        lines.append(line)
    return "\n".join(lines)

# Obligations exact fields (from DB)
OBL_SCALAR_KEYS = [
    "product", "bank_name", "bankName", "tenure", "roi",
    "total_loan", "totalLoan", "outstanding", "emi",
    "action", "transfer_to_proposed_bank", "existing_emi",
    "foirEmi", "selectedPercentage", "selectedTenurePercentage",
    "selectedRoiPercentage", "selected_percentage",
    "selected_tenure_percentage", "selected_roi_percentage",
]

# Applicant form fields (all from DB)
APPLICANT_FORM_KEYS = [
    "referenceNameForLogin", "aadharNumber", "qualification",
    "customerName", "panCard", "salaryAccountBank", "salaryAccountBankNumber",
    "mobileNumber", "alternateNumber", "fathersName", "ifscCode",
    "mothersName", "maritalStatus", "spousesName", "spousesDob",
    "currentAddress", "currentAddressLandmark", "currentAddressType",
    "currentAddressProof", "yearsAtCurrentAddress", "yearsInCurrentCity",
    "permanentAddress", "permanentAddressLandmark",
    "companyName", "yourDesignation", "yourDepartment", "dojCurrentCompany",
    "currentWorkExperience", "totalWorkExperience",
    "personalEmail", "workEmail", "officeAddress", "officeAddressLandmark",
    "ref1Name", "ref1Mobile", "ref1Relation", "ref1Address",
    "ref2Name", "ref2Mobile", "ref2Relation", "ref2Address",
    "formSubmittedAt",
]

# ─── STYLING ───────────────────────────────────────────────────────────────────

def style_header(ws, hex_color="1F4E79"):
    fill  = PatternFill("solid", fgColor=hex_color)
    font  = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font; cell.alignment = align

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, min_w), max_w)

def zebra(ws, c1="EBF5FB", c2="FDFEFE"):
    f1 = PatternFill("solid", fgColor=c1)
    f2 = PatternFill("solid", fgColor=c2)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        f = f1 if i % 2 == 0 else f2
        for cell in row:
            cell.fill = f
            cell.alignment = Alignment(vertical="center")

# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("📡  Connecting …")
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=15000)
    db = client[DB_NAME]

    # User + dept maps
    user_map, dept_map = {}, {}
    try:
        for u in db["users"].find({}, {"_id":1,"full_name":1,"name":1,"username":1,"email":1}):
            user_map[str(u["_id"])] = u.get("full_name") or u.get("name") or u.get("username") or u.get("email") or str(u["_id"])
    except: pass
    try:
        for d in db["departments"].find({}, {"_id":1,"name":1}):
            dept_map[str(d["_id"])] = d.get("name", str(d["_id"]))
    except: pass

    print("📥  Fetching leads …")
    leads = list(db[COLLECTION].find({}).sort("created_at", pymongo.DESCENDING))
    total = len(leads)
    print(f"  ✔  {total} leads")

    # ── Pass 1: collect all headers ───────────────────────────────────────────
    all_lead_cols   = {}   # ordered dict: col_name → None
    all_obl_cols    = {}
    all_apform_cols = {}

    for lead in leads:
        # Top-level
        top_flat = {}
        for k, v in lead.items():
            if k in SKIP_TOP:
                continue
            if isinstance(v, dict):
                top_flat.update(flatten(v, prefix=k))
            elif isinstance(v, list):
                if all(not isinstance(i, (dict, list)) for i in v):
                    top_flat[k] = ", ".join(to_str(i) for i in v)
            else:
                top_flat[k] = to_str(v)

        # dynamic_fields (excluding skip list)
        dyn = lead.get("dynamic_fields") or {}
        if isinstance(dyn, dict):
            for k, v in dyn.items():
                if k in SKIP_DYN:
                    continue
                fk = f"dynamic_fields.{k}"
                if isinstance(v, dict):
                    top_flat.update(flatten(v, prefix=fk))
                elif isinstance(v, list):
                    if all(not isinstance(i, (dict, list)) for i in v):
                        top_flat[fk] = ", ".join(to_str(i) for i in v)
                else:
                    top_flat[fk] = to_str(v)

        for col in top_flat:
            all_lead_cols[col] = None

        # Obligations
        obligations = (dyn.get("obligations") or []) if isinstance(dyn, dict) else []
        for ob in obligations:
            if isinstance(ob, dict):
                for k in ob:
                    all_obl_cols[k] = None

        # Applicant form
        af = (dyn.get("applicant_form") or {}) if isinstance(dyn, dict) else {}
        if isinstance(af, dict):
            for k in af:
                all_apform_cols[k] = None

    lead_headers = list(all_lead_cols.keys())
    obl_extra    = [k for k in all_obl_cols if k not in OBL_SCALAR_KEYS]
    obl_all_keys = OBL_SCALAR_KEYS + obl_extra
    apform_keys  = APPLICANT_FORM_KEYS + [k for k in all_apform_cols if k not in APPLICANT_FORM_KEYS]

    # Add applicant_form columns into leads
    apform_lead_cols  = [f"dynamic_fields.applicant_form.{k}" for k in apform_keys]
    # Insert obligations column right after the standard lead fields, before applicant_form
    OBL_SUMMARY_COL = "obligations"
    all_final_headers = lead_headers + [OBL_SUMMARY_COL] + [c for c in apform_lead_cols if c not in lead_headers]

    print(f"  Leads sheet cols  : {len(all_final_headers)}")
    print(f"  Obligations cols  : {len(obl_all_keys) + 5}")

    # ── Pass 2: build rows ────────────────────────────────────────────────────
    wb          = openpyxl.Workbook()
    ws_leads    = wb.active
    ws_leads.title = "Leads"
    ws_obl      = wb.create_sheet("Obligations")
    ws_stats    = wb.create_sheet("Stats")

    ws_leads.append(all_final_headers)

    # Obligations sheet: friendly header names (screenshot format)
    obl_friendly = {
        "product":                   "Product Type",
        "bank_name":                 "Bank Name",
        "bankName":                  "bankName (alt)",
        "tenure":                    "Tenure (months)",
        "roi":                       "ROI (%)",
        "total_loan":                "Total Loan",
        "totalLoan":                 "totalLoan (alt)",
        "outstanding":               "Outstanding",
        "emi":                       "EMI",
        "action":                    "Action",
        "transfer_to_proposed_bank": "Transfer to Proposed Bank",
        "existing_emi":              "Existing EMI",
        "foirEmi":                   "FOIR EMI",
        "selectedPercentage":        "Selected %",
        "selectedTenurePercentage":  "Selected Tenure %",
        "selectedRoiPercentage":     "Selected ROI %",
        "selected_percentage":       "selected_percentage",
        "selected_tenure_percentage":"selected_tenure_percentage",
        "selected_roi_percentage":   "selected_roi_percentage",
    }
    obl_display = [obl_friendly.get(k, k) for k in obl_all_keys]
    obl_headers = ["#", "Lead ID", "Lead Name", "Phone", "Status"] + obl_display
    ws_obl.append(obl_headers)

    obl_count = 0

    for idx, lead in enumerate(leads, start=1):
        # ── Build flat row ────────────────────────────────────────────────────
        flat = {}
        for k, v in lead.items():
            if k in SKIP_TOP:
                continue
            if isinstance(v, dict):
                flat.update(flatten(v, prefix=k))
            elif isinstance(v, list):
                if all(not isinstance(i, (dict, list)) for i in v):
                    flat[k] = ", ".join(to_str(i) for i in v)
            else:
                flat[k] = to_str(v)

        dyn = lead.get("dynamic_fields") or {}
        if isinstance(dyn, dict):
            for k, v in dyn.items():
                if k in SKIP_DYN:
                    continue
                fk = f"dynamic_fields.{k}"
                if isinstance(v, dict):
                    flat.update(flatten(v, prefix=fk))
                elif isinstance(v, list):
                    if all(not isinstance(i, (dict, list)) for i in v):
                        flat[fk] = ", ".join(to_str(i) for i in v)
                else:
                    flat[fk] = to_str(v)

        # Applicant form
        af = (dyn.get("applicant_form") or {}) if isinstance(dyn, dict) else {}
        if isinstance(af, dict):
            for k, v in af.items():
                flat[f"dynamic_fields.applicant_form.{k}"] = to_str(v)

        # Obligations summary (formatted, for Leads sheet)
        obligations = (dyn.get("obligations") or []) if isinstance(dyn, dict) else []
        flat[OBL_SUMMARY_COL] = fmt_obligations(obligations)

        # Resolve names
        assigned = lead.get("assigned_to")
        if isinstance(assigned, list):
            flat["assigned_to"] = ", ".join(user_map.get(str(a), str(a)) for a in assigned if a)
        elif assigned:
            flat["assigned_to"] = user_map.get(str(assigned), str(assigned))

        ar = lead.get("assign_report_to")
        if isinstance(ar, list):
            flat["assign_report_to"] = ", ".join(user_map.get(str(a), str(a)) for a in ar if a)

        if lead.get("created_by"):
            flat["created_by"] = user_map.get(str(lead["created_by"]), str(lead["created_by"]))

        if lead.get("department_id"):
            flat["department_id"] = dept_map.get(str(lead["department_id"]), str(lead["department_id"]))

        ws_leads.append([flat.get(h, "") for h in all_final_headers])

        # ── Obligation rows ───────────────────────────────────────────────────
        lead_id   = to_str(lead.get("custom_lead_id") or lead.get("_id"))
        full_name = f"{to_str(lead.get('first_name'))} {to_str(lead.get('last_name'))}".strip()
        phone     = to_str(lead.get("phone") or lead.get("mobile_number"))
        status    = to_str(lead.get("status"))

        for ob_no, ob in enumerate(obligations, start=1):
            if not isinstance(ob, dict): continue
            obl_row = [ob_no, lead_id, full_name, phone, status]
            for key in obl_all_keys:
                obl_row.append(to_str(ob.get(key)))
            ws_obl.append(obl_row)
            obl_count += 1

        if idx % 500 == 0 or idx == total:
            print(f"  … {idx}/{total}")

    # ── Style sheets ─────────────────────────────────────────────────────────
    style_header(ws_leads)
    zebra(ws_leads)
    auto_width(ws_leads)
    ws_leads.freeze_panes = "A2"

    # obligations column — wrap text, wider, top-aligned
    if OBL_SUMMARY_COL in all_final_headers:
        obl_col_idx = all_final_headers.index(OBL_SUMMARY_COL) + 1
        obl_col_letter = get_column_letter(obl_col_idx)
        ws_leads.column_dimensions[obl_col_letter].width = 65
        obl_fill  = PatternFill("solid", fgColor="FFF9E6")
        obl_font  = Font(size=9)
        obl_align = Alignment(vertical="top", wrap_text=True)
        # header cell — keep dark blue, no wrap override
        for row_cells in ws_leads.iter_rows(min_row=2, min_col=obl_col_idx, max_col=obl_col_idx):
            for cell in row_cells:
                cell.fill      = obl_fill
                cell.font      = obl_font
                cell.alignment = obl_align
        # Set row height for rows with multiple obligations
        for row_idx in range(2, ws_leads.max_row + 1):
            cell_val = ws_leads.cell(row=row_idx, column=obl_col_idx).value or ""
            lines    = cell_val.count("\n") + 1
            ws_leads.row_dimensions[row_idx].height = max(15, lines * 15)

    style_header(ws_obl, hex_color="145A32")
    zebra(ws_obl, c1="EAFAF1", c2="FDFEFE")
    auto_width(ws_obl)
    ws_obl.freeze_panes = "A2"

    # ── Stats ─────────────────────────────────────────────────────────────────
    def breakdown(data, title):
        counts = {}
        for v in data:
            counts[v or "Unknown"] = counts.get(v or "Unknown", 0) + 1
        ws_stats.append([f"── {title} ──", ""])
        ws_stats.append(["Value", "Count"])
        for v, c in sorted(counts.items(), key=lambda x: -x[1]):
            ws_stats.append([v, c])
        ws_stats.append([])

    ws_stats.append(["Metric", "Value"])
    ws_stats.append(["Total Leads", total])
    ws_stats.append(["Total Obligation Rows", obl_count])
    ws_stats.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_stats.append([])

    breakdown([l.get("status") for l in leads], "Status")
    breakdown([l.get("loan_type") or l.get("loan_type_name") for l in leads], "Loan Type")

    style_header(ws_stats, hex_color="6C3483")
    auto_width(ws_stats, max_w=60)

    # ── Save ─────────────────────────────────────────────────────────────────
    print(f"\n💾  Saving → {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    client.close()

    kb = os.path.getsize(OUTPUT_FILE) / 1024
    print(f"\n✅  Done!")
    print(f"    File       : {OUTPUT_FILE}")
    print(f"    Size       : {kb:.1f} KB ({kb/1024:.2f} MB)")
    print(f"    Leads      : {total}")
    print(f"    Lead cols  : {len(all_final_headers)}")
    print(f"    Obligations: {obl_count} rows")

if __name__ == "__main__":
    main()
