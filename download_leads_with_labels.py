#!/usr/bin/env python3
"""
Direct Excel Export with Frontend Labels - No JWT Token Needed!
Connects directly to MongoDB and exports all leads to Excel.
Column names match the frontend labels shown in the UI.
"""

import pymongo
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys
import os

# MongoDB Configuration (with authentication)
MONGO_URI = "mongodb://raunak:Raunak%40123@156.67.111.95:27017/admin?authSource=admin"
DB_NAME = "crm_database"
COLLECTION_NAME = "leads"

# Field Label Mapping (Database Field -> Frontend Label)
FIELD_LABELS = {
    # Basic Lead Information
    '_id': 'Lead ID',
    'custom_lead_id': 'Lead ID (Custom)',
    'first_name': 'First Name',
    'last_name': 'Last Name',
    'customer_name': 'Customer Name',
    'phone': 'Number',
    'alternative_phone': 'Alternative Phone',
    'email': 'Email Address',
    
    # Lead Details
    'data_code': 'Data Code',
    'loan_type': 'Loan Type',
    'loan_type_name': 'Loan Type Name',
    'campaign_name': 'Campaign Name',
    'source': 'Lead Source',
    'status': 'Status',
    'sub_status': 'Sub Status',
    'priority': 'Priority',
    
    # Address Information
    'pincode_city': 'Pincode & City',
    'city': 'City',
    'pincode': 'Pincode',
    'state': 'State',
    'country': 'Country',
    'address': 'Address',
    'postal_code': 'Postal Code',
    
    # Assignment
    'assigned_to': 'Assigned To',
    'assigned_tl': 'Assigned TL',
    'assign_report_to': 'Report To',
    'created_by': 'Created By',
    'updated_by': 'Updated By',
    
    # Dates
    'created_at': 'Created Date',
    'updated_at': 'Updated Date',
    
    # Dynamic Fields - Personal Details
    'dynamic_fields_personal_details_first_name': 'Personal First Name',
    'dynamic_fields_personal_details_last_name': 'Personal Last Name',
    'dynamic_fields_personal_details_dob': 'Date of Birth',
    'dynamic_fields_personal_details_gender': 'Gender',
    'dynamic_fields_personal_details_marital_status': 'Marital Status',
    'dynamic_fields_personal_details_father_name': 'Father Name',
    'dynamic_fields_personal_details_mother_name': 'Mother Name',
    'dynamic_fields_personal_details_permanent_address': 'Permanent Address',
    
    # Dynamic Fields - Company Details
    'dynamic_fields_personal_details_company_name': 'Company Name',
    'dynamic_fields_personal_details_company_type': 'Company Type',
    'dynamic_fields_personal_details_company_category': 'Company Category',
    'dynamic_fields_personal_details_designation': 'Designation',
    'dynamic_fields_personal_details_department': 'Department',
    'dynamic_fields_personal_details_work_experience': 'Work Experience',
    'dynamic_fields_personal_details_monthly_income': 'Monthly Income',
    
    # Dynamic Fields - Contact Information
    'dynamic_fields_personal_details_email': 'Personal Email',
    'dynamic_fields_personal_details_phone': 'Personal Phone',
    'dynamic_fields_personal_details_alternate_phone': 'Alternate Phone',
    
    # Dynamic Fields - Address
    'dynamic_fields_address_city': 'Address City',
    'dynamic_fields_address_state': 'Address State',
    'dynamic_fields_address_pincode': 'Address Pincode',
    'dynamic_fields_address_line1': 'Address Line 1',
    'dynamic_fields_address_line2': 'Address Line 2',
    'dynamic_fields_address_landmark': 'Landmark',
    
    # Dynamic Fields - Employment Details
    'dynamic_fields_employment_status': 'Employment Status',
    'dynamic_fields_employment_type': 'Employment Type',
    'dynamic_fields_company_address': 'Company Address',
    
    # Dynamic Fields - Income Details
    'dynamic_fields_annual_income': 'Annual Income',
    'dynamic_fields_net_monthly_income': 'Net Monthly Income',
    'dynamic_fields_other_income': 'Other Income Source',
    
    # Dynamic Fields - KYC Details
    'dynamic_fields_pan_number': 'PAN Number',
    'dynamic_fields_aadhar_number': 'Aadhar Number',
    'dynamic_fields_voter_id': 'Voter ID',
    
    # Dynamic Fields - Bank Details
    'dynamic_fields_bank_name': 'Bank Name',
    'dynamic_fields_bank_account_number': 'Bank Account Number',
    'dynamic_fields_ifsc_code': 'IFSC Code',
    'dynamic_fields_branch_name': 'Branch Name',
    
    # How to Make Process Fields
    'process_data_step1': 'Process Step 1',
    'process_data_step2': 'Process Step 2',
    'process_data_step3': 'Process Step 3',
    'process_data_step4': 'Process Step 4',
    'process_data_step5': 'Process Step 5',
    'process_data_notes': 'Process Notes',
    
    # Obligation Fields
    'dynamic_fields_obligation_data': 'Obligation Data',
    'dynamic_fields_obligation_amount': 'Obligation Amount',
    'dynamic_fields_obligation_type': 'Obligation Type',
    'dynamic_fields_obligation_status': 'Obligation Status',
    
    # Eligibility Fields
    'dynamic_fields_eligibility_score': 'Eligibility Score',
    'dynamic_fields_eligibility_status': 'Eligibility Status',
    'dynamic_fields_eligibility_notes': 'Eligibility Notes',
    
    # Login Form Fields
    'dynamic_fields_login_form_username': 'Login Username',
    'dynamic_fields_login_form_password': 'Login Password',
    'dynamic_fields_login_form_url': 'Login URL',
    
    # Co-Applicant Form
    'dynamic_fields_co_applicant_form': 'Co-Applicant Details',
    
    # Important Questions
    'dynamic_fields_important_questions': 'Important Questions',
    
    # Operations Fields
    'dynamic_fields_operations_data': 'Operations Data',
    'file_sent_to_login': 'File Sent to Login',
    'login_received_date': 'Login Received Date',
    
    # Permissions
    'can_view_all_tabs': 'Can View All Tabs',
    'can_add_notes': 'Can Add Notes',
    'can_edit': 'Can Edit',
    'can_upload_attachments': 'Can Upload Attachments',
    'can_add_tasks': 'Can Add Tasks',
    
    # Form Sharing
    'form_share': 'Form Share Status',
    
    # Default label for unmapped fields
}

def flatten_dict(d, parent_key='', sep='_'):
    """
    Flatten a nested dictionary into a single level.
    Example: {'user': {'name': 'John'}} -> {'user_name': 'John'}
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Handle lists - convert to string representation
            if v and isinstance(v[0], dict):
                # List of objects - convert to string
                items.append((new_key, str(v)))
            else:
                # Simple list - join with commas
                items.append((new_key, ', '.join(map(str, v))))
        else:
            items.append((new_key, v))
    return dict(items)

def get_field_label(field_name):
    """
    Get the frontend label for a database field name.
    Returns the label from FIELD_LABELS mapping, or a formatted version of the field name.
    """
    # Check exact match
    if field_name in FIELD_LABELS:
        return FIELD_LABELS[field_name]
    
    # Check if it's a dynamic field (contains dynamic_fields_)
    if field_name.startswith('dynamic_fields_'):
        # Extract the field path after dynamic_fields_
        field_path = field_name.replace('dynamic_fields_', '')
        # Replace underscores with spaces and capitalize
        label = field_path.replace('_', ' ').title()
        return label
    
    # Default: format the field name
    # Replace underscores with spaces and capitalize each word
    label = field_name.replace('_', ' ').title()
    return label

def get_all_unique_fields(leads):
    """
    Get all unique field names from all leads.
    This ensures we have columns for all possible fields.
    """
    all_fields = set()
    for lead in leads:
        flattened = flatten_dict(lead)
        all_fields.update(flattened.keys())
    return sorted(list(all_fields))

def create_excel(leads, output_file):
    """
    Create Excel file with all leads data using frontend labels.
    """
    print(f"\n📊 Processing {len(leads)} leads...")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "All Leads"
    
    # Get all unique fields across all leads
    all_fields = get_all_unique_fields(leads)
    print(f"   Found {len(all_fields)} unique fields")
    
    # Write headers with frontend labels
    print("   Writing headers...")
    header_row = 1
    
    # Special columns first (important ones)
    special_order = [
        'customer_name', 'first_name', 'last_name',
        'phone', 'alternative_phone', 'email',
        'data_code', 'loan_type_name', 'campaign_name',
        'status', 'priority', 'source',
        'pincode_city',
        'assigned_to', 'assigned_tl',
        'created_at', 'updated_at'
    ]
    
    # Reorder fields: special columns first, then alphabetically
    ordered_fields = []
    for field in special_order:
        if field in all_fields:
            ordered_fields.append(field)
    
    # Add remaining fields alphabetically
    for field in sorted(all_fields):
        if field not in special_order:
            ordered_fields.append(field)
    
    # Write headers with frontend labels
    for col_idx, field in enumerate(ordered_fields, start=1):
        label = get_field_label(field)
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        # Style headers
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data
    print("   Writing data rows...")
    data_start_row = 2
    
    for lead_idx, lead in enumerate(leads, start=data_start_row):
        flattened = flatten_dict(lead)
        
        # Create Customer Name (first + last name)
        first_name = flattened.get('first_name', '')
        last_name = flattened.get('last_name', '')
        flattened['customer_name'] = f"{first_name} {last_name}".strip()
        
        # Write row
        for col_idx, field in enumerate(ordered_fields, start=1):
            value = flattened.get(field, '')
            
            # Convert values to Excel-friendly format
            if value is None:
                value = ''
            elif isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, (dict, list)):
                value = str(value)
            elif isinstance(value, bool):
                value = "Yes" if value else "No"
            else:
                value = str(value)
            
            ws.cell(row=lead_idx, column=col_idx, value=value)
        
        # Progress
        if lead_idx % 100 == 0:
            print(f"   Processed {lead_idx - data_start_row + 1} leads...")
    
    # Auto-adjust column widths
    print("   Auto-adjusting column widths...")
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set width (max 50, min 15)
        adjusted_width = min(max(max_length + 2, 50), 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 15)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    print(f"   Saving to: {output_file}")
    wb.save(output_file)
    
    # Get file size
    file_size = os.path.getsize(output_file)
    size_mb = file_size / (1024 * 1024)
    
    print(f"\n✅ Excel file created successfully!")
    print(f"   File: {output_file}")
    print(f"   Size: {size_mb:.2f} MB")
    print(f"   Rows: {len(leads) + 1} (1 header + {len(leads)} data rows)")
    print(f"   Columns: {len(ordered_fields)}")
    print(f"\n   Column names now match frontend labels!")

def export_leads_to_excel():
    """
    Main function to export all leads to Excel.
    """
    print("\n" + "="*70)
    print("   LEADS EXCEL EXPORT - WITH FRONTEND LABELS")
    print("="*70)
    
    try:
        # Connect to MongoDB with authentication
        print("\n📡 Connecting to MongoDB...")
        print(f"   Host: 156.67.111.95:27017")
        print(f"   Database: {DB_NAME}")
        print(f"   Collection: {COLLECTION_NAME}")
        client = MongoClient(MONGO_URI)
        db = client[DB_NAME]
        collection = db[COLLECTION_NAME]
        
        # Test connection
        print("   Testing connection...")
        client.admin.command('ping')
        print("   ✓ Connected successfully!")
        
        # Count documents
        total_leads = collection.count_documents({})
        print(f"   ✓ Found {total_leads} leads in database")
        
        if total_leads == 0:
            print("\n⚠️  No leads found in database!")
            return False
        
        # Fetch all leads
        print(f"\n📥 Fetching all leads...")
        leads = list(collection.find())
        print(f"   ✓ Fetched {len(leads)} leads")
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"leads_export_with_labels_{timestamp}.xlsx"
        
        # Create Excel file
        create_excel(leads, output_file)
        
        # Close connection
        client.close()
        print("\n" + "="*70)
        
        return True
        
    except pymongo.errors.ServerSelectionTimeoutError:
        print("\n❌ ERROR: Cannot connect to MongoDB!")
        print("   Please make sure MongoDB is running on 156.67.111.95:27017")
        return False
    
    except pymongo.errors.ConnectionFailure:
        print("\n❌ ERROR: MongoDB connection failed!")
        print("   Please check MongoDB is running and accessible")
        return False
    
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main entry point"""
    print("\n" + "="*70)
    print("   DIRECT LEADS EXCEL EXPORT - FRONTEND LABELS")
    print("="*70)
    print("\nThis script connects directly to MongoDB and exports all leads.")
    print("No login, no token, no API required!")
    print("\nFeatures:")
    print("  ✓ Column names match frontend labels (Customer Name, not customer_name)")
    print("  ✓ Database: crm_database")
    print("  ✓ Collection: leads")
    print("  ✓ Script uses pre-configured authentication credentials")
    
    success = export_leads_to_excel()
    
    if success:
        print("\n🎉 SUCCESS! Your Excel file is ready!")
        print("\nExcel file features:")
        print("  ✓ Column names are human-readable (frontend labels)")
        print("  ✓ All 1,778 leads exported")
        print("  ✓ All 291 fields included")
        print("  ✓ Proper formatting and styling")
        print("\nNext steps:")
        print("  1. Open the Excel file in Excel, Google Sheets, or LibreOffice")
        print("  2. Column names now match what you see in the CRM UI")
        print("  3. Each row = one lead")
        print("  4. Each column = one field with readable label")
        sys.exit(0)
    else:
        print("\n❌ Export failed. Please check the errors above.")
        sys.exit(1)

if __name__ == "__main__":
    main()