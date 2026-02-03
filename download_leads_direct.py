#!/usr/bin/env python3
"""
Direct Excel Export - No JWT Token Needed!
Connects directly to MongoDB and exports all leads to Excel.
"""

import pymongo
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys
import os

# MongoDB Configuration (with authentication)
MONGO_URI = "mongodb://raunak:Raunak%40123@156.67.111.95:27017/admin?authSource=admin"
DB_NAME = "crm_database"
COLLECTION_NAME = "leads"

def flatten_dict(d, parent_key='', sep='_'):
    """
    Flatten a nested dictionary into a single level.
    Example: {'user': {'name': 'John'}} -> {'user_name': 'John'}
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Handle lists - convert to string representation
            if v and isinstance(v[0], dict):
                # List of objects - convert to string
                items.append((new_key, str(v)))
            else:
                # Simple list - join with commas
                items.append((new_key, ', '.join(map(str, v))))
        else:
            items.append((new_key, v))
    return dict(items)

def get_all_unique_fields(leads):
    """
    Get all unique field names from all leads.
    This ensures we have columns for all possible fields.
    """
    all_fields = set()
    for lead in leads:
        flattened = flatten_dict(lead)
        all_fields.update(flattened.keys())
    return sorted(list(all_fields))

def create_excel(leads, output_file):
    """
    Create Excel file with all leads data.
    """
    print(f"\n📊 Processing {len(leads)} leads...")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "All Leads"
    
    # Get all unique fields across all leads
    all_fields = get_all_unique_fields(leads)
    print(f"   Found {len(all_fields)} unique fields")
    
    # Write headers
    print("   Writing headers...")
    header_row = 1
    
    # Special columns first
    special_order = [
        'first_name', 'last_name', 'customer_name',
        'status', 'priority', 'lead_source',
        'created_at', 'updated_at'
    ]
    
    # Reorder fields: special columns first, then alphabetically
    ordered_fields = []
    for field in special_order:
        if field in all_fields:
            ordered_fields.append(field)
    
    # Add remaining fields alphabetically
    for field in sorted(all_fields):
        if field not in special_order:
            ordered_fields.append(field)
    
    # Write headers
    for col_idx, field in enumerate(ordered_fields, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=field)
        # Style headers
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data
    print("   Writing data rows...")
    data_start_row = 2
    
    for lead_idx, lead in enumerate(leads, start=data_start_row):
        flattened = flatten_dict(lead)
        
        # Create Customer Name (first + last name)
        first_name = flattened.get('first_name', '')
        last_name = flattened.get('last_name', '')
        flattened['customer_name'] = f"{first_name} {last_name}".strip()
        
        # Write row
        for col_idx, field in enumerate(ordered_fields, start=1):
            value = flattened.get(field, '')
            
            # Convert values to Excel-friendly format
            if value is None:
                value = ''
            elif isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, (dict, list)):
                value = str(value)
            elif isinstance(value, bool):
                value = "Yes" if value else "No"
            else:
                value = str(value)
            
            ws.cell(row=lead_idx, column=col_idx, value=value)
        
        # Progress
        if lead_idx % 100 == 0:
            print(f"   Processed {lead_idx - data_start_row + 1} leads...")
    
    # Auto-adjust column widths
    print("   Auto-adjusting column widths...")
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set width (max 50, min 15)
        adjusted_width = min(max(max_length + 2, 50), 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 15)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    print(f"   Saving to: {output_file}")
    wb.save(output_file)
    
    # Get file size
    file_size = os.path.getsize(output_file)
    size_mb = file_size / (1024 * 1024)
    
    print(f"\n✅ Excel file created successfully!")
    print(f"   File: {output_file}")
    print(f"   Size: {size_mb:.2f} MB")
    print(f"   Rows: {len(leads) + 1} (1 header + {len(leads)} data rows)")
    print(f"   Columns: {len(ordered_fields)}")

def export_leads_to_excel():
    """
    Main function to export all leads to Excel.
    """
    print("\n" + "="*70)
    print("   LEADS EXCEL EXPORT - DIRECT FROM MONGODB")
    print("="*70)
    
    try:
        # Connect to MongoDB with authentication
        print("\n📡 Connecting to MongoDB...")
        print(f"   Host: 156.67.111.95:27017")
        print(f"   Database: {DB_NAME}")
        print(f"   Collection: {COLLECTION_NAME}")
        client = MongoClient(MONGO_URI)
        db = client[DB_NAME]
        collection = db[COLLECTION_NAME]
        
        # Test connection
        print("   Testing connection...")
        client.admin.command('ping')
        print("   ✓ Connected successfully!")
        
        # Count documents
        total_leads = collection.count_documents({})
        print(f"   ✓ Found {total_leads} leads in database")
        
        if total_leads == 0:
            print("\n⚠️  No leads found in database!")
            return False
        
        # Fetch all leads
        print(f"\n📥 Fetching all leads...")
        leads = list(collection.find())
        print(f"   ✓ Fetched {len(leads)} leads")
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"leads_export_{timestamp}.xlsx"
        
        # Create Excel file
        create_excel(leads, output_file)
        
        # Close connection
        client.close()
        print("\n" + "="*70)
        
        return True
        
    except pymongo.errors.ServerSelectionTimeoutError:
        print("\n❌ ERROR: Cannot connect to MongoDB!")
        print("   Please make sure MongoDB is running on localhost:27017")
        return False
    
    except pymongo.errors.ConnectionFailure:
        print("\n❌ ERROR: MongoDB connection failed!")
        print("   Please check MongoDB is running and accessible")
        return False
    
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main entry point"""
    print("\n" + "="*70)
    print("   DIRECT LEADS EXCEL EXPORT - NO AUTHENTICATION NEEDED!")
    print("="*70)
    print("\nThis script connects directly to MongoDB and exports all leads.")
    print("No login, no token, no API required!")
    print("\nRequirements:")
    print("  ✓ MongoDB must be running on 156.67.111.95:27017")
    print("  ✓ Database name: crm_database")
    print("  ✓ Collection: leads")
    print("  ✓ Script uses pre-configured authentication credentials")
    
    success = export_leads_to_excel()
    
    if success:
        print("\n🎉 SUCCESS! Your Excel file is ready!")
        print("\nNext steps:")
        print("  1. Open the Excel file in Excel, Google Sheets, or LibreOffice")
        print("  2. All leads from your CRM are in the file")
        print("  3. Each row = one lead")
        print("  4. Each column = one field")
        sys.exit(0)
    else:
        print("\n❌ Export failed. Please check the errors above.")
        sys.exit(1)

if __name__ == "__main__":
    main()